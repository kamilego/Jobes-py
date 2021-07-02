from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
from math import sqrt
import time
import codecs
import csv


begin_time = time.time()


def check_value(row):
    z = max(l[3] for l in row)
    k = min(l[3] for l in row)
    if abs(k) < abs(z):
        return str(row[0][2])+"_"+str(row[0][0]), round(z,2)
    else:
        return str(row[0][2])+"_"+str(row[0][0]), round(k,2)


def change(elem):
    if elem.count("/") == 3:
        pos = elem.index('/', elem.index('/') + 1)
        elem = elem[:pos] + 'KAM' + elem[pos+1:]
    return elem


def cell_properties(cell,color):
    cell.fill = PatternFill(fgColor=color, fill_type="solid")
    thin = Side(border_style="thin", color="000000")
    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def cells_range(rowik,a,b):
    sum_val = 0
    for num in range(a,b):
        sum_val += (ws.cell(rowik, num).value)**2
    return round(sqrt(sum_val),2)


def sumq(row, value):
    ws.cell(row, value).value = cells_range(row, value+1, value+komb_d)
    cell_properties(ws.cell(row, value),"00FF6600")


def minus_value_check(row, col):
    if ws.cell(row,col+1).value < 0:
        return ws.cell(row,col+1).value - ws.cell(row+1,col+1).value
    else:
        return ws.cell(row,col+1).value + ws.cell(row+1,col+1).value


def list_check(lis):
    if len([elem for elem in lis if elem < 0]) == 9:
        return lis[::-1]
    else:
        return lis


def max_force(row,col,a):
    val_list = []
    for num in range(3,len(kombinacje),komb_d):
        val_list.append(minus_value_check(row-1,num))
    val_list = sorted(val_list)
    val_list = list_check(val_list)
    if a == 1:  
        value = val_list[-1]
    else:
        value = val_list[0]
    ws.cell(row, col).value = value
    cell_properties(ws.cell(row, col),"00008000")


def cell_values(num1, step, row):
    for num in range(num1, num1+step):
        ws.cell(row+2, num).value = ws.cell(row+1, num1-1).value - ws.cell(row+1, num).value


def cell_value_fill(row, num, name, color):
    ws.cell(row, num).value = name
    cell_properties(ws.cell(row, num), color)


path = r"D:\kamil\_scripts\pythony\siedlce\krawezniki\krawezniki.csv"
division = 25
komb_d = 11

path2 = path[:-3] + "xlsx"
path3 = path[:-4] + "_temp.xlsx"

with codecs.open(path,"rb","utf-16") as f:
    a = csv.reader(f,delimiter='\t')
    a = [''.join(elem) for elem in list(a)[1:]]

for i in range(len(a)):
    a[i] = change(a[i])
    a[i] = a[i].replace("(K)","").replace("\n","").replace("/",";").replace("KAM","/").split(";")
a = [row[:4] for row in a]

for num in range(len(a)):
    for i in range(2):
        a[num][i] = a[num][i].strip()
    a[num][0] = int(a[num][0])
    a[num][2] = int(a[num][2])
    a[num][3] = float(a[num][3].replace(",","."))

a = sorted(a, key=lambda x: x[2])
all_list = [a[num:num+division] for num in range(0,len(a),division)]
diction = {}

for row in all_list:
    diction[check_value(row)[0]] = check_value(row)[1]

aa = {}
rod_num = int(list(diction.keys())[-1].split("_")[1])-int(list(diction.keys())[0].split("_")[1]) + 1

for elem in range(rod_num):
	for num in range(0,len(diction),rod_num):
		aa[list(diction.keys())[elem+num]] = diction[list(diction.keys())[elem+num]]

prety = sorted(set([int(elem.split("_")[1]) for elem in list(aa.keys())]))
kombinacje = sorted(set([int(elem.split("_")[0]) for elem in list(aa.keys())]))

wb = load_workbook(path3)
ws = wb.active

row1 = 3
num1 = 3
for name in prety:
    cell_value_fill(row1, num1, name, "00C0C0C0")
    row1 += 3

for row in range(2,len(prety)*3,3):
    for num, name in enumerate(kombinacje,start=4):
        cell_value_fill(row, num, name, "00666699")

ami = 0
for to in range(3,len(prety)*3+1,3):
    for ro in range(4,len(kombinacje)+4):
        cell_value_fill(to, ro, list(aa.values())[ami], "00FFCC99")
        ami+=1

for to in range(2,len(prety)*3+2,3):
    for num in range(5,len(kombinacje)-1,komb_d):
        cell_values(num, komb_d-1, to)
    for num in range(4,len(kombinacje)+3,komb_d):
        sumq(to+2,num)
    max_force(to+2, 2, 0)
    max_force(to+2, 1, 1)


wb.save(path2)


print(round(time.time() - begin_time,2))

