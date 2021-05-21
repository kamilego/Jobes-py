from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
from math import sqrt
import time
import codecs
import csv

begin_time = time.time()

def check_value(row):
    z = max(l[3] for l in row)
    k = min(l[3] for l in row)
    if abs(k) < abs(z):
        return str(row[0][2])+"_"+str(row[0][0]), round(z,2)
    else:
        return str(row[0][2])+"_"+str(row[0][0]), round(k,2)


def change(elem):
    if elem.count("/") == 3:
        pos = elem.index('/', elem.index('/') + 1)
        elem = elem[:pos] + 'KAM' + elem[pos + 1:]
    return elem


def cell_properties(cell,color):
    y = cell
    y.fill = PatternFill(fgColor=color, fill_type = "solid")
    thin = Side(border_style="thin", color="000000")
    y.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def cells_range(rowik, a,b):
    sum_val = 0
    for num in range(a,b):
        sum_val += (ws.cell(rowik, num).value)**2
    return round(sqrt(sum_val),2)


def sumq(row, value):
    ws.cell(row, value).value = cells_range(row, value+1, value+13)
    cell_properties(ws.cell(row, value),"00FF6600")


def minus_value_check(row, col):
    if ws.cell(row,col).value < 0:
        return ws.cell(row,col).value - ws.cell(row+1,col).value
    else:
        return ws.cell(row,col).value + ws.cell(row+1,col).value


def list_check(lis):
    if len([elem for elem in lis if elem < 0]) == 3:
        return lis[::-1]
    else:
        return lis


def max_force(row,col, a):
    val_list = sorted([minus_value_check(row-1,3),minus_value_check(row-1,16), minus_value_check(row-1,29)])
    val_list = list_check(val_list)
    if a == 1:  
        value = val_list[-1]
    else:
        value = val_list[0]
    ws.cell(row, col).value = value
    cell_properties(ws.cell(row, col),"00008000")


path = r"D:\kamil\RTON Kisielice\_NOWE\wykrat\wykrat3.csv"
path2 = path[:-3] + "xlsx"

with codecs.open(path,"rb","utf-16") as f:
    a = csv.reader(f,delimiter='\t')
    a = [''.join(elem) for elem in list(a)[1:]]

for i in range(len(a)):
    a[i] = change(a[i])
    a[i] = a[i].replace("(K)","").replace("\n","").replace("/",";").replace("KAM","/").split(";")

test = [a[i:975+i] for i in range(0,len(a),975)]

for i in range(len(test)):
    test[i] = sorted(test[i], key=lambda x: x[2])
    for elem in test[i]:
        elem[0] = int(elem[0].strip())
        elem[2] = int(elem[2])
        elem[3] = float(elem[3].replace(",","."))

all_list = []

for elem in test:
    for num in range(0,975,25):
        all_list.append(elem[num:num+25])

diction ={}

for row in all_list:
    diction[check_value(row)[0]] = check_value(row)[1]

prety = sorted(set([int(elem.split("_")[1]) for elem in list(diction.keys())]))
kombinacje = sorted(set([int(elem.split("_")[0]) for elem in list(diction.keys())]))

wb = Workbook()
ws = wb.active

row1 = 3
num1 = 2
for name in prety:
    ws.cell(row1, num1).value = name
    cell_properties(ws.cell(row1, num1),"00C0C0C0")
    row1 += 3


for ami in range(0,len(diction.values()),39):
    for to in range(3,len(prety)*3+1,3):
        for ro in range(3,len(kombinacje)+3):
            if to == int(ami/13)+3:
                ws.cell(to, ro).value = list(diction.values())[ro+ami-3]
                cell_properties(ws.cell(to, ro),"00FFCC99")
            
for to in range(2,len(prety)*3+2,3):
    for ro in range(3,len(kombinacje)+3):
        ws.cell(to, ro).value = (list(diction.keys())[ro-3]).split("_")[0]
        cell_properties(ws.cell(to, ro),"00666699")
    for num in range(4,16):
        ws.cell(to+2, num).value = ws.cell(to+1, 3).value - ws.cell(to+1, num).value
    for num in range(17,29):
        ws.cell(to+2, num).value = ws.cell(to+1, 16).value - ws.cell(to+1, num).value
    for num in range(30,42):
        ws.cell(to+2, num).value = ws.cell(to+1, 29).value - ws.cell(to+1, num).value
    for num in range(3,30,13):
        sumq(to+2,num)
    max_force(to+2, 2, 1)
    max_force(to+2, 1, 0)

list_of_values = []
min_list_of_values = []

def max_list(num, col, a):
    max_li = []
    for i in range(num,num+23*3,3):
        max_li.append(ws.cell(i,col).value)
    if a == 1:
        return max(max_li)
    else:
        return min(max_li)


ws2 = wb.create_sheet("Segmenty")
for num in range(4,len(prety)*3+2,72):
    list_of_values.append(max_list(num, 2, 1))
    min_list_of_values.append(max_list(num, 1, 0))

elem_max = max(list_of_values)

for row in range(2, len(list_of_values)+3):
    for col in range(8,9):
        if row == 2:
            ws2.cell(row,6).value = "Minusy"
            ws2.cell(row,7).value = "Segment"
            ws2.cell(row,8).value = "Wartość"
            ws2.cell(row,9).value = "Wytężenie"
        else:
            ws2.cell(row,6).value = min_list_of_values[row-3]
            ws2.cell(row,7).value = row-2
            ws2.cell(row,8).value = list_of_values[row-3]
            cell_properties(ws2.cell(row, col),"00FFFFFF")
            if row-3 == list_of_values.index(elem_max):
                cell_properties(ws2.cell(row, 8),"00FF6600")

nb14_2 = 217.70
nb12_5 = 217.70

for row in range(3, len(list_of_values)+3):
    if row == 8 or row == 14 or row == 20 or row == 26:
        ws2.cell(row, 9).value = list_of_values[row-3]/nb14_2
        ws2.cell(row, 10).value = str(round((list_of_values[row-3]/nb14_2)*100))+"%"
    else:
        ws2.cell(row, 9).value = list_of_values[row-3]/nb12_5
        ws2.cell(row, 10).value = str(round((list_of_values[row-3]/nb12_5)*100))+"%"

wb.save(path2)


print(round(time.time() - begin_time,2))
