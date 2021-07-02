import codecs
import csv
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side


begin_time = time.time()


def cell_properties(cell,color):
    cell.fill = PatternFill(fgColor=color, fill_type = "solid")
    thin = Side(border_style="thin", color="000000")
    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


path = "D:\kamil\_scripts\pythony\siedlce\przemieszczenia\przemieszczenia.csv"
komb_d = 11

path2 = path[:-3] + "xlsx"
path3 = path[:-4] + "_temp.xlsx"


with codecs.open(path,"rb","utf-16") as f:
    results = csv.reader(f,delimiter='\t')
    results = [elem for elem in results][1:]

new_results = []

for elem in results:
    for i in range(len(elem)):
        elem[i] = elem[i].replace("\n", "").replace(",", ".").replace("/", ";").split(";")
        for num in range(3):
            if not num == 2:
                elem[i][num] = int(elem[i][num].strip())
            else:
                elem[i][num] = float(elem[i][num])
        new_results.append(elem[i])

results = [new_results[i:komb_d+i] for i in range(0, len(new_results), komb_d)]
elem_names = set([elem[num][0] for elem in results for num in range(len(elem))])
komb_num = set([elem[num][1] for elem in results for num in range(len(elem))])


wb = load_workbook(path3)
ws = wb.active

row1 = 3
num1 = 2
for name in elem_names:
    ws.cell(row1, num1).value = name
    cell_properties(ws.cell(row1, num1),"00C0C0C0")
    row1 += 3

step = 0

for row in range(2, len(elem_names)*3, 3):
    for num in range(3, len(komb_num)+3):
        ws.cell(row, num).value = komb_num[num-3]
        ws.cell(row+1, num).value = new_results[step][2]
        step += 1


wb.save(path2)


print(round(time.time() - begin_time,2))

