from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
import time
import codecs
import csv


begin_time = time.time()


def change(elem):
    if elem.count("/") == 3:
        pos = elem.index('/', elem.index('/') + 1)
        elem = elem[:pos] + 'KAM' + elem[pos + 1:]
    return elem


def cell_properties(cell,color):
    cell.fill = PatternFill(fgColor=color, fill_type="solid")
    thin = Side(border_style="thin", color="000000")
    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


path = r"D:\kamil\_scripts\pythony - Kopia\siedlce\reakcje\fundamenty.csv"
division = 25
komb_d = 11

path2 = path[:-3] + "xlsx"
path3 = path[:-4] + "_temp.xlsx"

with codecs.open(path,"rb","utf-16") as f:
    a = csv.reader(f,delimiter='\t')
    a = [''.join(elem) for elem in list(a)[1:]]

for i in range(len(a)):
    a[i] = change(a[i])
    a[i] = a[i].replace("(K)","").replace("\n","").replace("/",";").replace("KAM","/").split(";")

a = a[:[a.index(elem) for elem in a if " " in elem][0]]

test = [a[i:komb_d+i] for i in range(0,len(a),komb_d)]

for i in range(len(test)):
    for elem in test[i]:
        if elem[2] == '-0,000':
            elem[2] = int(0)
        else:
            elem[2] = float(elem[2].replace(",","."))
        elem[0] = int(elem[0].strip())
        elem[1] = int(elem[1])
        for num in range(3,5):
            elem[num] = float(elem[num].replace(",","."))
    test[i] = sorted(test[i], key=lambda x: x[1])

komb_set = set([row[1] for elem in test for row in elem])

index_kom_set = [num for num in range(4,int(len(komb_set)+len(komb_set)/komb_d-1+4))]
ab = [index_kom_set[num:num+komb_d] for num in range(0, len(komb_set), komb_d)]

for step, elem in enumerate(ab[1:], start=1):
	for num in range(len(elem)):
		elem[num] = elem[num] + step

num_of_elem = set([elem[0] for elem in a])
indexes = [num for elem in ab for num in elem]
komb_index = [(x,y) for x,y in zip(indexes,komb_set)]
fx = [elem[2] for num in range(len(test)) for elem in test[num]]
fy = [elem[3] for num in range(len(test)) for elem in test[num]]
fz = [elem[4] for num in range(len(test)) for elem in test[num]]
col_indexs = [num for elem in [indexes]*len(num_of_elem) for num in elem]
row_ixdes = [[num]*len(komb_set) for num in range(3, len(num_of_elem)*9, 9)]
row_ixdes = [num for elem in row_ixdes for num in elem]
most_values = [(x,y,z,k,row) for x,y,z,k,row in zip(fx,fy,fz,col_indexs,row_ixdes)]


wb = load_workbook(path3)
ws = wb.active

row1 = 2
for name in range(1, len(num_of_elem)+1):
    ws.cell(row1, 3).value = name
    cell_properties(ws.cell(row1, 3),"00C0C0C0")
    row1 += 9

for num in range(2, len(num_of_elem)*9, 9):
    ws.cell(num,1).value = "FX"
    ws.cell(num+3,1).value = "FY"
    ws.cell(num+6,1).value = "FZ"

for row in range(2, len(num_of_elem)*9, 3):
    for num, elem in komb_index:
        ws.cell(row, num).value = elem

for number in range(3):
    for elem in most_values:
        ws.cell(elem[4]+number*3, elem[3]).value = elem[number]


wb.save(path2)


print(round(time.time() - begin_time,2))

