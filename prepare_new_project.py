import os
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill
from openpyxl.styles import colors
from openpyxl.styles import Border, Side, Font, Alignment
import shutil
import datetime
import re


class NewProject:
    def __init__(self, new_project, new_date, new_rev_1, new_rev_2, lok1, lok2):
        self.new_project = new_project
        self.new_date = new_date
        self.new_rev_1 = new_rev_1
        self.new_rev_2 = new_rev_2
        self.lok1 = lok1
        self.lok2 = lok2
        self.newpath = '%s\%s' % (r'D:\kamil', self.new_project)
        self.dic_w_replace = {'GDP0004B': self.new_project,
                              '210402': self.new_date , 
                              "REW.01": self.new_rev_1, 
                              "REW.02": self.new_rev_2}


    def first_steps(self):
        '''Function that creates new project folder and copy there three files with modified filed only
           If fles and folders already exists there will be special comment'''
        while True:
            # creating new folder for new project and paste there example .dwg and xlsx
            print_path = self.newpath + "\pdf"
            if not os.path.exists(self.newpath):
                os.makedirs(self.newpath)
                os.makedirs(print_path)
                shutil.copy(r"D:\kamil\_scripts\templates\Drawing5.dwg","%(x)s\%(y)s\%(y)s.dwg" % {"x": r'D:\kamil', "y": self.new_project})
                shutil.copy(r"D:\kamil\_scripts\templates\KALKULATOR_KK_nieghaslo.xlsx","%(x)s\%(y)s\%(y)s.xlsx" % {"x": r'D:\kamil', "y": self.new_project})
            else:
                print("New project path already exists.")
                break
            # opening example file and replace values for ploting script
            with open(r"D:\kamil\_scripts\templates\template.dsd") as b:
                a = b.readlines()
                for i in range(len(a)):
                    for key, value in self.dic_w_replace.items():
                        a[i] = a[i].replace(key, value)

            # creating .dsd file - plotting file
            with open("%(x)s\%(y)s\%(y)s.dsd" % {"x": r'D:\kamil', "y": self.new_project}, "w") as f:
                f.writelines("%s" % elem for elem in a)

            # creating .scr file - execution script file
            with open("%(x)s\%(y)s\%(y)s.scr" % {"x": r'D:\kamil', "y": self.new_project}, "w") as s:
                s.writelines('-PUBLISH\n%(x)s/%(y)s/%(y)s.dsd\n' % {"x": 'D:/kamil', "y": self.new_project})

            # opening script file to change tabs layout dwg names
            with open(r"D:\kamil\_scripts\templates\uklady.scr") as k:
                l = k.readlines()
                for i in range(3,24,4):
                    for key, value in self.dic_w_replace.items():
                        l[i] = l[i].replace(key, value)

            # creating .scr file - changing tabs names
            with open("%(x)s\%(y)s\%(y)suklad.scr" % {"x": r'D:\kamil', "y": self.new_project}, "w") as g:
                g.writelines("%s" % elem for elem in l)
        
            print("Everything has been copied to new project path.")
            break


    def create_excel_histry(self):
        '''Function that creates new excel with actual hyperlink to folders
           If xlsx file exists previous is copied to _old folder
           All situations were checked due that there will be special comment about states'''
        lista = os.listdir(r'D:\kamil')
        del lista[0:2]
        del lista[-4:-1]
        if os.path.exists("D:\kamil\historia.xlsx"):
            lista.remove("historia.xlsx")
        lista = lista[0:len(lista)-1]
        now = str(datetime.datetime.now())[:19]
        now = now.replace(":","_").replace(" ","___")
        while True:
            if os.path.exists("D:\kamil\historia.xlsx"):
                while True:
                    wb = load_workbook("D:\kamil\historia.xlsx")
                    ws = wb.active
                    length = len(lista)
                    first_set = set()
                    second_set = set(lista)
                    for num in range(length):
                        first_set.add(ws["A"+str(num+1)].value)
                    if self.new_project in first_set:
                        print("Project already exists in excel list.")
                        break
                    if len(first_set - second_set) == 0:
                        print("List of folders in excel and path are the same\nNothing has been added.")
                        break
                    shutil.copy("D:\kamil\historia.xlsx", os.path.join("D:\kamil\_old\historia_" + now + ".xlsx"))
                    row_num = lista.index(self.new_project)+2
                    ws.insert_rows(row_num, amount=1)
                    thin = Side(border_style="thin", color="000000")
                    for row in ws["A%s:S%s" % (row_num, row_num)]:
                        for cell in row:
                            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    second_set = list(second_set)
                    second_set.sort()
                    for num, elem in enumerate(second_set, start=2):
                        ws['A'+str(num)].hyperlink = os.path.join("D:/kamil/", elem)
                        ws['A'+str(num)].value = elem
                        ws['A'+str(num)].style = "Hyperlink"
                    self.new_date = self.new_date[:2] + "-" + self.new_date[2:4] + "-" + self.new_date[4:]
                    ws['B'+str(row_num)] = self.new_date
                    work_sheet_a1 = ws['B'+str(row_num)]
                    work_sheet_a1.font = Font(color='000000', name="Consolas")
                    work_sheet_a1.number_format = '0'
                    wb.save("D:\kamil\historia.xlsx")
                    print("Added new folder hyperlink path to excel.")
                    break
                break
            wb = Workbook()
            ws = wb.active
            ws.column_dimensions["A"].width = 12
            for num, elem in enumerate(lista, start=2):
                wb.active['A'+str(num)].hyperlink = "D:/kamil/" + elem
                wb.active['A'+str(num)].value = elem
                wb.active['A'+str(num)].style = "Hyperlink"
            wb.save("D:\kamil\historia.xlsx")
            print("Excel haven't existed. Created excel file with hyperlinks paths.")
            break


    def check_platform(self, tower):
        if "E2" in tower:
            return "PRM-B2.35"
        elif "E3" in tower:
            return "PRM-B3.35"
        else:
            return "PRM-B4.35"


    def execution(self):
        '''Creates script to change information in AutoCAD about new towers localization'''
        while True:
            if not os.path.exists(self.newpath):
                print("New project path doesnt exist. Script wasn't copied")
                break
            elif os.path.exists(self.newpath+r"\tagi.scr"):
                print("Scipt already exists in folder.")
                break
            else:
                proj_name = self.new_project[:3]+" "+self.new_project[3:7]+" "+self.new_project[7:]
                path = r"D:\kamil\_scripts\templates\tagi.SCR"
                date = datetime.datetime.now().isoformat()[:10]
                date = date[5:7]+"."+date[:4]
                area = self.lok2[self.lok2.index(".",self.lok2.index(".")+1)+1:]
                tower_type = re.compile(r'E\d/\d\d').findall(self.lok1)[0]
                platform_type = NewProject.check_platform(self, tower_type)
                list_values = [proj_name, date, self.lok1, self.lok2, area, tower_type, platform_type]

                with open(path) as b:
                    a = b.readlines()
                num_list = [elem for elem in range(10,len(a),12)]
                dictionary = {key: value for key, value in zip(num_list, list_values)}
                for key, value in dictionary.items():
                    a[key] = value + "\n"

                with open(os.path.join(r'D:\kamil', self.new_project, "tagi.scr"), "w") as g:
                    g.writelines("%s" % elem for elem in a)
                print("Scipt has been copied.")
                break
