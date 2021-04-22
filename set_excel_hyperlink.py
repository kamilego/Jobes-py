import os
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import shutil
import datetime


# replaces second appearing word to destiny word 
def replace_second(string, a, b): 
   return string.replace(a, b, 2).replace(b, a, 1) 


# function that creates new folder of project and copy there three files with modified filed only
def first_steps(base_path, new_project, dwg_dir, xlsx_path, dic_w_replace):
    while True:
        # creating new folder for new project and paste there example .dwg and xlsx
        newpath = '%s\%s' % (base_path, new_project)
        print_path = newpath + "\pdf"
        if not os.path.exists(newpath):
            os.makedirs(newpath)
            os.makedirs(print_path)
            shutil.copy(dwg_dir,"%(x)s\%(y)s\%(y)s.dwg" % {"x": base_path, "y": new_project})
            shutil.copy(xlsx_path,"%(x)s\%(y)s\%(y)s.xlsx" % {"x": base_path, "y": new_project})
        else:
            print("New project path already exists.\nNothing has been done.")
            break

        # opening example file and replace values for ploting script
        b = open(r"D:\kamil\_scripts\templates\template.dsd")
        a = b.readlines()
        for i in range(len(a)):
            for key, value in dic_w_replace.items():
                a[i] = a[i].replace(key, value)

        # creating .dsd file - plotting file
        f = open("%(x)s\%(y)s\%(y)s.dsd" % {"x": base_path, "y": new_project}, "w")
        with f as filehandle:
            filehandle.writelines("%s" % elem for elem in a)

        # creating .scr file - execution script file
        s = open("%(x)s\%(y)s\%(y)s.scr" % {"x": base_path, "y": new_project}, "w")
        with s as filehandle:
            filehandle.writelines('-PUBLISH\n%(x)s/%(y)s/%(y)s.dsd\n' % {"x": base_path.replace("\\","/"), "y": new_project})

        # opening script file to change tabs layout dwg names
        k = open(r"D:\kamil\_scripts\templates\uklady.scr")
        l = k.readlines()
        for i in range(3,24,4):
            for key, value in dic_w_replace.items():
                l[i] = l[i].replace(key, value)

        # creating .scr file - changing tabs names
        g = open("%(x)s\%(y)s\%(y)suklad.scr" % {"x": base_path, "y": new_project}, "w")
        with g as filehandle:
            filehandle.writelines("%s" % elem for elem in l)
    
        print("Everything has been copied to new project path.")
        break


# function that compare two lists, first list of existing folders and list in excel with folders
# if lists are the same program returns information about that
# if not it returns information that all was added
def edit_excel_histry(list_values):
    while True:
        wb = load_workbook("D:\kamil\historia.xlsx")
        ws = wb.active
        length = len(list_values)
        first_set = set()
        second_set = set(list_values)
        for num in range(length):
            first_set.add(ws["A"+str(num+1)].value)
        if len(first_set - second_set) == 0:
            print("List of folders in excel and path are the same\nNothing has been added.")
            break
        row_num = list_values.index(list(second_set - first_set)[0])+1
        ws.insert_rows(row_num)
        second_set = list(second_set)
        second_set.sort()
        for num, elem in enumerate(second_set, start=1):
            wb.active['A'+str(num)].hyperlink = "D:/kamil/"+elem
            wb.active['A'+str(num)].value = elem
            wb.active['A'+str(num)].style = "Hyperlink"
        wb.save("D:\kamil\historia.xlsx")
        print("Added new folder hyperlink path to excel.")
        break


# function that reads actual list of folders in path and returns list of them
def list_of_folders():
    lista = (os.listdir('D:/kamil/'))
    del lista[0:2]
    del lista[-4:-1]
    if os.path.exists("D:\kamil\historia.xlsx"):
        lista.remove("historia.xlsx")
    return lista[0:len(lista)-1]


# function that creates new excel with actual hyperlink to folders
# if xlsx file exists previous is copied to _old folder
# use it only in critical case
def create_excel_histry(lista):
    now = str(datetime.datetime.now())[:19]
    now = now.replace(":","_").replace(" ","___")
    while True:
        if os.path.exists("D:\kamil\historia.xlsx"):
            print("All is fine. Excel file already exists.")
            shutil.copy("D:\kamil\historia.xlsx", "D:\kamil\_old\historia_" + now + ".xlsx")
            break
        wb = Workbook()
        ws = wb.active
        ws.column_dimensions["A"].width = 12
        for num, elem in enumerate(lista, start=2):
            wb.active['A'+str(num)].hyperlink = "D:/kamil/"+elem
            wb.active['A'+str(num)].value = elem
            wb.active['A'+str(num)].style = "Hyperlink"
        wb.save("D:\kamil\historia.xlsx")
        print("Created excel file with hyperlinks paths.")
        break
