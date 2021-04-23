import os
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook


# modify string output of data about modified files [rr-mm-dd]
def modification_date(filename):
    temp = os.path.getmtime(filename)
    string = str(datetime.datetime.fromtimestamp(temp))
    string = string[2:10]
    return string


# creates list of folders from path and modify it
def list_of_folders(path):
    fold_list = os.listdir(path)
    fold_list = fold_list[2:-4]
    fold_list.remove('historia.xlsx')
    return fold_list


# renames folders names from "PDF" to "pdf"
def rename_folders_name(path):
    fold_list = list_of_folders(path)
    sub_folders = [path + "\\" + x for x in fold_list]
    for i in range(len(sub_folders)):
        for elem in os.listdir(sub_folders[i]):
            if "PDF" == elem:
                os.rename(sub_folders[i] + "\PDF", sub_folders[i] + "\pdf")


# returns dictinary with project names as keys and string about modify data as values
def folders_mod_date(path):
    pdfs_folders = ["%s\%s\pdf" % (path, x) for x in list_of_folders(path)]
    pdfs_paths = {}
    for i in range(len(pdfs_folders)):
        for elem in os.listdir(pdfs_folders[i]):
            if "PS01" in elem:
                pdfs_paths[pdfs_folders[i][9:17]] = modification_date("%s\%s" % (pdfs_folders[i], elem))
    return pdfs_paths


# creates excel with kayes and values from dictionary
def excel_with_names(pdfs_dictionary):
    wb = Workbook()
    ws = wb.active
    for i, (k, v) in enumerate(pdfs_dictionary.items(), start=5):
        ws['D'+str(i)] = k
        ws['E'+str(i)] = v
    wb.save(r"C:\Users\Proj-North\Desktop\OLD\asd.xlsx")


path = "D:\kamil"
new_pdfs_paths = folders_mod_date(path)
# a = list_of_folders(path)
# rename_folders_name(path)
# excel_with_names(new_pdfs_paths)
